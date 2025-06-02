可運作

import streamlit as st
import datetime
import pandas as pd
from io import BytesIO
import calendar
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ===== 主日數與幸運物件資料 =====
day_meaning = {
    1: {"名稱": "創造日", "指引": "展現創意，展現自我魅力。", "星": "⭐⭐⭐⭐"},
    2: {"名稱": "連結日", "指引": "適合合作，溝通與等待機會。", "星": "⭐⭐"},
    3: {"名稱": "表達日", "指引": "表達想法，展現自我魅力。", "星": "⭐⭐⭐"},
    4: {"名稱": "實作日", "指引": "建立基礎，適合細節與規劃。", "星": "⭐⭐⭐"},
    5: {"名稱": "行動日", "指引": "啟動新的計畫，做出主動選擇。", "星": "⭐⭐⭐⭐"},
    6: {"名稱": "關係日", "指引": "接觸愛情，適當調整。", "星": "⭐⭐⭐"},
    7: {"名稱": "內省日", "指引": "適合學習、休息與自我對話。", "星": "⭐"},
    8: {"名稱": "成果日", "指引": "聚焦目標與務成就。", "星": "⭐⭐⭐⭐"},
    9: {"名稱": "釋放日", "指引": "放手，療癒與完成階段。", "星": "⭐⭐"},
}

lucky_map = {
    1: {"色": "🔴 紅色", "水晶": "紅瑪瑙", "小物": "原子筆"},
    2: {"色": "🟠 橘色", "水晶": "太陽石", "小物": "月亮吊飾"},
    3: {"色": "🟡 黃色", "水晶": "黃水晶", "小物": "紙膠帶"},
    4: {"色": "🟢 綠色", "水晶": "綠幽靈", "小物": "方形石頭"},
    5: {"色": "🔵 淺藍色", "水晶": "拉利瑪", "小物": "交通票卡"},
    6: {"色": "🔷 靛色", "水晶": "青金石", "小物": "愛心吊飾"},
    7: {"色": "🟣 紫色", "水晶": "紫水晶", "小物": "書籤"},
    8: {"色": "💗 粉色", "水晶": "粉晶", "小物": "鋼筆"},
    9: {"色": "⚪ 白色", "水晶": "白水晶", "小物": "小香包"},
}

# ===== 工具函式 =====
def reduce_to_digit(n):
    while n > 9:
        n = sum(int(x) for x in str(n))
    return n

def format_layers(total):
    mid = sum(int(x) for x in str(total))
    return f"{total}/{mid}/{reduce_to_digit(mid)}" if mid > 9 else f"{total}/{mid}"

def get_flowing_year_ref(query_date, bday):
    query_date = query_date.date() if hasattr(query_date, "date") else query_date
    cutoff = datetime.date(query_date.year, bday.month, bday.day)
    return query_date.year - 1 if query_date < cutoff else query_date.year

def get_flowing_month_ref(query_date, birthday):
    query_date = query_date.date() if hasattr(query_date, "date") else query_date
    if query_date.day < birthday.day:
        return query_date.month - 1 if query_date.month > 1 else 12
    return query_date.month

def get_additional_guidance(flowing_day):
    main_number = reduce_to_digit(flowing_day)
    if main_number == 5:
        if flowing_day == 32:
            return "這一天，創意與行動的平衡將帶來新的計畫，準備好啟動變革。"
        elif flowing_day == 41:
            return "這一天，務實的行動將與創意結合，為新機會打下基礎。"
    elif main_number == 1:
        return "今天是展示創意與自我的好時機，讓你吸引更多的目光與機會。"
    elif main_number == 2:
        return "今天是適合合作與溝通的日子，耐心等待機會的來臨。"
    elif main_number == 3:
        return "自信表達自己的想法，與他人分享你的創意與理念。"
    elif main_number == 4:
        return "這一天是規劃與執行的最佳時機，專注細節並做好準備。"
    elif main_number == 6:
        return "關注他人需求，今天是營造和諧關係的日子。"
    elif main_number == 7:
        return "給自己一些安靜的時間，進行深層的內省與學習。"
    elif main_number == 8:
        return "聚焦於目標，今天是行動的最佳時機，邁向成就。"
    elif main_number == 9:
        return "放下過去，準備迎接新的階段，療癒自己。"
    return ""

def style_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="流年月曆")
        workbook = writer.book
        worksheet = workbook["流年月曆"]
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        for idx, column in enumerate(df.columns):
            worksheet.column_dimensions[chr(65 + idx)].width = 15
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in worksheet.iter_rows():
            worksheet.row_dimensions[row[0].row].height = 35
    return output

st.set_page_config(page_title="樂覺製所生命靈數", layout="centered")
st.title("🧭 樂覺製所生命靈數")
st.markdown("在數字之中，\n我們與自己不期而遇。\n**Be true, be you — 讓靈魂，自在呼吸。**")

birthday = st.date_input("請輸入生日", value=datetime.date(1990, 1, 1), min_value=datetime.date(1900, 1, 1))
target_year = st.number_input("請選擇年份", min_value=1900, max_value=2100, value=datetime.datetime.now().year)
target_month = st.selectbox("請選擇月份", list(range(1, 13)), index=datetime.datetime.now().month - 1)

if st.button("🎉 產生日曆建議表"):
    _, last_day = calendar.monthrange(target_year, target_month)
    days = pd.date_range(start=datetime.date(target_year, target_month, 1),
                         end=datetime.date(target_year, target_month, last_day))
    data = []
    for d in days:
        fd_total = sum(int(x) for x in f"{birthday.year}{birthday.month:02}{d.day:02}")
        flowing_day = format_layers(fd_total)
        main_number = reduce_to_digit(fd_total)
        meaning = day_meaning.get(main_number, {})
        lucky = lucky_map.get(main_number, {})
        additional_guidance = get_additional_guidance(fd_total)
        guidance = meaning.get("指引", "") + " " + additional_guidance
        year_ref = get_flowing_year_ref(d, birthday)
        fy_total = sum(int(x) for x in f"{year_ref}{birthday.month:02}{birthday.day:02}")
        flowing_year = format_layers(fy_total)
        fm_ref = get_flowing_month_ref(d, birthday)
        fm_total = sum(int(x) for x in f"{birthday.year}{fm_ref:02}{birthday.day:02}")
        flowing_month = format_layers(fm_total)
        date_str = d.strftime("%Y-%m-%d")
        weekday_str = d.strftime("%A")
        data.append({
            "日期": date_str,
            "星期": weekday_str,
            "流年": flowing_year,
            "流月": flowing_month,
            "流日": flowing_day,
            "運勢指數": meaning.get("星", ""),
            "指引": guidance,
            "幸運色": lucky.get("色", ""),
            "水晶": lucky.get("水晶", ""),
            "幸運小物": lucky.get("小物", "")
        })
    df = pd.DataFrame(data)
    st.dataframe(df)
    file_name = f"LuckyCalendar_{target_year}_{str(target_month).zfill(2)}.xlsx"
    title = "樂覺製所生命靈數"
    subtitle = "在數字之中，我們與自己不期而遇。Be true, be you — 讓靈魂，自在呼吸。"
    if not df.empty and df.dropna(how='all').shape[0] > 0:
        output = style_excel(df)
        st.markdown(f"### {title}")
        st.markdown(f"**{subtitle}**")
        st.download_button(
            "📥 點此下載 " + file_name.replace(".xlsx", " 年靈數流日建議表（三層加總斜線版）"),
            data=output.getvalue(),
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.warning("⚠️ 無法匯出 Excel：目前資料為空，請先產生日曆資料")
